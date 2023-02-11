from .models import *
from .serializers import *
import os
import json
import requests
import openpyxl
from datetime import datetime
import urllib.request
import re
from Accounts.models import *
from Accounts.serializer import *

from rest_framework.authtoken.models import Token
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.authentication import TokenAuthentication
from rest_framework import status

from django.core.files.storage import default_storage
from django.conf import settings



def fetch_latest_news():
    print("Fetching Latest News, migrating database.")

    url = '''https://newsapi.org/v2/everything?q="farmer"  and "plants" and "india" &from=2023-01-11&sortBy=publishedAt&apiKey=cd97861cf26741f19e1e4a27d2310f9d&searchIn
    =title,description&language=en'''
    x = requests.get(url)
    data = json.loads(x.text)
    
    keywords = ['india', 'farmer', 'garden', 'plant']

    News.objects.all().delete()

    for i in data["articles"]:
        flag = False
        for k in keywords: 
            if flag: break
            if (k in i['title'].lower() or k in i['description'].lower()) and 'india' in (i['title'].lower() or i['description'].lower()) :
                flag = True
                d = {
                    'source_name' : i['source']['name'], 
                    'author' : i['author'], 
                    'title' : i['title'], 
                    'description' : i['description'], 
                    'url' : i['url'], 
                    'urlToImage' : i['urlToImage'], 
                    'publishedAt' : i['publishedAt'], 
                    'content' : i['content']
                }
                news = News.objects.create(**d)
                news.save()
    
    print("Updated latest news")
                

    

from apscheduler.schedulers.background import BackgroundScheduler
scheduler = BackgroundScheduler()
scheduler.add_job(fetch_latest_news, 'interval', minutes = 10)
scheduler.start()




@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([AllowAny]) # IsAuthenticated
def index(request):
    data = {
        'message' : "Home Page for API's. Make sure not to mess around ;)"
    }
    return Response(data)
    



@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([AllowAny]) # IsAuthenticated
def schemes(request):
    items = Schemes.objects.order_by("pk")
    ser = SchemesSerializer(items, many=True)
    return Response(ser.data)
    

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([AllowAny]) # IsAuthenticated
def news(request):
    items = News.objects.order_by("pk")
    ser = NewsSerializer(items, many=True)
    return Response(ser.data, status=status.HTTP_200_OK)
    


@api_view(["POST"])
@authentication_classes([TokenAuthentication])
@permission_classes([AllowAny])
def upload(request):
    print('hi')
    f = request.FILES['image']
    # data = json.loads(request.body)
    # f = data['file']
    # print(f)
    file_name = default_storage.save(f.name, f)
    url = os.path.join(settings.MEDIA_URL,file_name)
    data = {
            'slug' : f'{url}',
            'type' : f.content_type
        }
    print(data)

    return Response(data)

@api_view(['GET','POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([AllowAny]) # IsAuthenticated
def feed(request):
    if request.method == 'GET':
        items = Feed.objects.order_by("-pk")
        username = request.GET.get('username', None)
        ser = FeedSerializer(items, many=True, context={'username': username})
        return Response(ser.data)
    elif request.method == 'POST':
        author = request.data['author']
        user_type = request.data['user_type']
        image = request.data['image']
        caption = request.data['caption']
        likes = 0
        feed = Feed.objects.create(author=author, user_type=user_type, image=image, caption=caption, likes=likes)
        feed.save()
        d = {
            'message':'success'
        }
        return Response(d)


    


@api_view(['GET','POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([AllowAny]) # IsAuthenticated
def chat(request,user1,user2):
    if request.method == 'GET':
        print(user1,user2)
        return Response(ser.data)
    

def extract_crops(temp, feels_like, humidity, path_crops):
    
    crops = []
    wb_obj = openpyxl.load_workbook(path_crops)
    sheet_obj = wb_obj.active

    max_rows = sheet_obj.max_row
    max_columns = sheet_obj.max_column

    for i in range(2,max_rows+1):

        temp_lower = float(sheet_obj.cell(row = i, column = 2).value)
        temp_upper = float(sheet_obj.cell(row = i, column = 3).value)
        humid = float(sheet_obj.cell(row = i, column = 6).value)

        if ((temp >= temp_lower and temp <= temp_upper) or (feels_like >= temp_lower and feels_like <= temp_upper)) and humidity <= humid:
            temporary = {
                'name' : sheet_obj.cell(row = i, column = 1).value,
                'soil' : sheet_obj.cell(row = i, column = 7).value,
                'temperature' : f'{temp_lower}-{temp_upper}',
                'rain' : f'{sheet_obj.cell(row = i, column = 4).value}-{sheet_obj.cell(row = i, column = 5).value}',
                'relative_humidity' : humid,
            }
            crops.append(temporary)

    return crops

def extract_plants(path_plants):
    wb_obj = openpyxl.load_workbook(path_plants)
    sheet_obj = wb_obj.active

    max_rows = sheet_obj.max_row
    max_columns = sheet_obj.max_column

    for i in range(2,max_rows+1):

        m = sheet_obj.cell(row = i, column = 1).value

        if m == datetime.now().strftime("%B").lower():
            plants = {
                'sow_indoors' : sheet_obj.cell(row = i, column = 2).value,
                'sow_outdoors' : sheet_obj.cell(row = i, column = 3).value,
                'harvest' : sheet_obj.cell(row = i, column = 4).value,
            }
            
            return plants

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([AllowAny]) # IsAuthenticated
def recommend(request):
    lat = request.GET.get('lat') 
    lon = request.GET.get('lon') 
    
    # To get Address 
    url = f'''https://api.geoapify.com/v1/geocode/reverse?lat={lat}&lon={lon}&apiKey=5ae5f7b11bfc4ad18ae8aed6e068408c'''
    x = requests.get(url)
    data = json.loads(x.text)
    location = data['features'][0]['properties']['formatted']

    # To get climate
    url = f'''https://api.openweathermap.org/data/2.5/onecall?lat={lat}&lon={lon}&appid=24646f23cc8d2aa7d34025558bc6d4cb&units=metric'''
    x = requests.get(url)
    data = json.loads(x.text)
    temperature = data['current']['temp']
    feels_like = data['current']['feels_like']
    humidity = data['current']['humidity']
    weather = data['current']['weather'][0]['main']

    # To get recommended Crops, Plants
    crops = extract_crops(temperature, feels_like, humidity, f'{settings.BASE_DIR}//custom_ds//crops.xlsx')
    plants = extract_plants(f'{settings.BASE_DIR}//custom_ds//plants.xlsx')



    response = {
        'location' : location,
        'climate' : {
            'temp' : temperature,
            'feels_like' : feels_like,
            'humidity' : humidity,
            'weather' : weather
        },
        'crops' : crops,
        'plants' : plants
    }
    return Response(response, status=status.HTTP_200_OK)

def get_title(VideoID):
    
    params = {"format": "json", "url": "https://www.youtube.com/watch?v=%s" % VideoID}
    url = "https://www.youtube.com/oembed"
    query_string = urllib.parse.urlencode(params)
    url = url + "?" + query_string

    with urllib.request.urlopen(url) as response:
        response_text = response.read()
        data = json.loads(response_text.decode())
        return data['title']

def scraper(search_keyword):
    
    html = urllib.request.urlopen(f"https://www.youtube.com/results?search_query={search_keyword}&sp=CAM%253D") # Filter for Most popular, Relevant and latest
    video_ids = re.findall(r"watch\?v=(\S{11})", html.read().decode())

    videos = []
    s = set()

    for i in range(0,len(video_ids)):
        if video_ids[i] not in s:
            temp = {
                'url' : f"https://www.youtube.com/watch?v={video_ids[i]}",
                'title' : get_title(video_ids[i]),
                'img1' : f'https://img.youtube.com/vi/{video_ids[i]}/0.jpg',
                'img2' : f'https://img.youtube.com/vi/{video_ids[i]}/1.jpg',
                'img3' : f'https://img.youtube.com/vi/{video_ids[i]}/2.jpg'
            }
            s.add(video_ids[i])
            videos.append(temp)
        if len(s) == 5:
            return videos

    return videos

    
@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([AllowAny]) # IsAuthenticated
def latest(request):

    # To get Scraped Latest Advancement videos
    videos = scraper('hybrid+farming+techniques')

    return Response(videos, status=status.HTTP_200_OK)
    

    
    
@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([AllowAny]) # IsAuthenticated
def like(request):
    username = request.GET.get('username') 
    post_id = request.GET.get('post_id')
    print(username)
    
    ins = isLiked.objects.create(username=username,post_id=post_id,liked=True)
    ins.save()

    feed = Feed.objects.get(id=post_id)
    feed.likes = feed.likes + 1
    feed.save()

    data = {
        'message' : 'success'
    }

    return Response(data)
    
    
@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([AllowAny]) # IsAuthenticated
def userlist(request):
    
    items = UserProfile.objects.order_by("pk")
    ser = UserProfileSerializer(items, many=True)
    return Response(ser.data)
    
    
@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([AllowAny]) # IsAuthenticated
def comments(request):
    print(request.data)
    author = request.data['author']
    comment = request.data['comment']
    post_id = request.data['post_id']

    ins = Comments.objects.create(author=author,comment=comment)
    ins.save()

    feed_ins = Feed.objects.get(id=post_id)
    feed_ins.comments.add(ins)
    feed_ins.save()
    
    d = {
        'message':'success'
    }

    return Response(d)
    
